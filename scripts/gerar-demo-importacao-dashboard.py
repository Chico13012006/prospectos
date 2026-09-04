from pathlib import Path
import csv
from openpyxl import Workbook
from openpyxl.chart import BarChart, DoughnutChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "demo"
OUT.mkdir(exist_ok=True)
XLSX = OUT / "prospectos-importacao-dashboard-demo.xlsx"
CSV = OUT / "prospectos-importacao-demo.csv"

headers = ["Nome", "E-mail", "Empresa", "Segmento", "Origem", "Telefone", "Cargo", "Cidade", "Estado"]
rows = [
    ["Ana Martins", "ana.martins@example.com", "Aurora Serviços", "Laudos Técnicos", "Prospecção própria", "11990001001", "Gerente Administrativa", "São Paulo", "SP"],
    ["Bruno Lima", "bruno.lima@example.com", "Beta Facilities", "Laudos Técnicos", "Indicação", "21990001002", "Diretor Operacional", "Rio de Janeiro", "RJ"],
    ["Carla Souza", "carla.souza@example.com", "Crescer Brinquedos", "Laudos Técnicos", "LinkedIn", "31990001003", "Proprietária", "Belo Horizonte", "MG"],
    ["Diego Alves", "diego.alves@example.com", "Delta Eventos", "Laudos Técnicos", "Prospecção própria", "41990001004", "Coordenador", "Curitiba", "PR"],
    ["Elisa Rocha", "elisa.rocha@example.com", "Estação Kids", "Laudos Técnicos", "Feira", "51990001005", "Sócia", "Porto Alegre", "RS"],
    ["Felipe Gomes", "felipe.gomes@example.com", "Futura Escolas", "Laudos Técnicos", "Google", "61990001006", "Comprador", "Brasília", "DF"],
    ["Gabriela Reis", "gabriela.reis@example.com", "Geração Lazer", "Laudos Técnicos", "Indicação", "71990001007", "Gerente", "Salvador", "BA"],
    ["Henrique Melo", "henrique.melo@example.com", "Horizonte Condomínios", "Laudos Técnicos", "Prospecção própria", "81990001008", "Síndico profissional", "Recife", "PE"],
    ["Isabela Costa", "isabela.costa@example.com", "Integra Play", "Laudos Técnicos", "Instagram", "85990001009", "Diretora", "Fortaleza", "CE"],
    ["João Ribeiro", "joao.ribeiro@example.com", "Jardim Encantado", "Laudos Técnicos", "Prospecção própria", "11990001010", "Administrador", "Campinas", "SP"],
    ["Karen Oliveira", "karen.oliveira@example.com", "Kairós Entretenimento", "Laudos Técnicos", "LinkedIn", "27990001011", "Coordenadora", "Vitória", "ES"],
    ["Lucas Nunes", "lucas.nunes@example.com", "Lúdica Parques", "Laudos Técnicos", "Indicação", "48990001012", "Sócio", "Florianópolis", "SC"],
]

with CSV.open("w", newline="", encoding="utf-8-sig") as f:
    writer = csv.writer(f, delimiter=";")
    writer.writerow(headers)
    writer.writerows(rows)

wb = Workbook()
ws = wb.active
ws.title = "IMPORTACAO_CSV"
dash = wb.create_sheet("DASHBOARD_DEMO")
guide = wb.create_sheet("LEIA-ME")

navy = "101522"
panel = "1A2133"
purple = "5B42F3"
green = "21C77A"
amber = "F2B84B"
slate = "94A3B8"
white = "F8FAFC"
border = Side(style="thin", color="30384D")

for sh in wb.worksheets:
    sh.sheet_view.showGridLines = False
    for row in sh.iter_rows():
        for cell in row:
            cell.font = Font(name="Arial", size=10)

# Import sheet
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:I{len(rows)+1}"
for c, value in enumerate(headers, 1):
    cell = ws.cell(1, c, value)
    cell.fill = PatternFill("solid", fgColor=purple)
    cell.font = Font(name="Arial", bold=True, color=white)
    cell.alignment = Alignment(vertical="center")
for r_idx, row in enumerate(rows, 2):
    for c_idx, value in enumerate(row, 1):
        cell = ws.cell(r_idx, c_idx, value)
        cell.font = Font(name="Arial", color="1E293B")
        cell.fill = PatternFill("solid", fgColor="FFFFFF" if r_idx % 2 == 0 else "F1F5F9")
        cell.border = Border(bottom=Side(style="hair", color="CBD5E1"))
ws.row_dimensions[1].height = 26
widths = [22, 32, 28, 22, 22, 18, 25, 20, 10]
for idx, width in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(idx)].width = width

# Dashboard source/status columns, fictitious operational state.
statuses = ["Novos Leads", "Novos Leads", "Primeiro Contato", "Aguardando Resposta", "Em Follow-up", "Respondeu", "Novos Leads", "Em Follow-up", "Reunião Agendada", "Aguardando Resposta", "Respondeu", "Novos Leads"]
dash.sheet_properties.pageSetUpPr.fitToPage = True
dash.sheet_view.zoomScale = 90
dash.merge_cells("A1:H2")
dash["A1"] = "ProspectOS · Dashboard demonstrativo"
dash["A1"].font = Font(name="Arial", bold=True, size=22, color=white)
dash["A1"].fill = PatternFill("solid", fgColor=navy)
dash["A1"].alignment = Alignment(vertical="center")
dash.merge_cells("A3:H3")
dash["A3"] = "Dados 100% fictícios · Exemplo visual da carteira após importação e ativação pelo gestor"
dash["A3"].font = Font(name="Arial", italic=True, color=slate)
dash["A3"].fill = PatternFill("solid", fgColor=navy)

cards = [("A5", "Total de leads", "=COUNTA(IMPORTACAO_CSV!A2:A1000)", purple),
         ("C5", "Responderam", '=COUNTIF(J12:J23,"Respondeu")', green),
         ("E5", "Em follow-up", '=COUNTIF(J12:J23,"Em Follow-up")', amber),
         ("G5", "Reuniões", '=COUNTIF(J12:J23,"Reunião Agendada")', "38BDF8")]
for anchor, label, formula, color in cards:
    col = dash[anchor].column
    dash.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col+1)
    dash.merge_cells(start_row=6, start_column=col, end_row=7, end_column=col+1)
    dash.cell(5, col, label).font = Font(name="Arial", bold=True, color=slate)
    dash.cell(6, col, formula).font = Font(name="Arial", bold=True, size=20, color=white)
    for rr in range(5, 8):
        for cc in range(col, col+2):
            dash.cell(rr, cc).fill = PatternFill("solid", fgColor=panel)
            dash.cell(rr, cc).border = Border(left=border, right=border, top=border, bottom=border)
    dash.cell(6, col).alignment = Alignment(vertical="center")
    dash.cell(6, col+1).fill = PatternFill("solid", fgColor=color)

dash["A9"] = "Carteira por etapa"
dash["A9"].font = Font(name="Arial", bold=True, size=14, color=white)
dash["A9"].fill = PatternFill("solid", fgColor=navy)
stage_names = ["Novos Leads", "Primeiro Contato", "Aguardando Resposta", "Em Follow-up", "Respondeu", "Reunião Agendada"]
for idx, stage in enumerate(stage_names, 10):
    dash.cell(idx, 1, stage)
    dash.cell(idx, 2, f'=COUNTIF($J$12:$J$23,A{idx})')
    dash.cell(idx, 1).font = Font(name="Arial", color=white)
    dash.cell(idx, 2).font = Font(name="Arial", bold=True, color=white)
    dash.cell(idx, 1).fill = dash.cell(idx, 2).fill = PatternFill("solid", fgColor=panel)

dash["D9"] = "Leads por UF"
dash["D9"].font = Font(name="Arial", bold=True, size=14, color=white)
dash["D9"].fill = PatternFill("solid", fgColor=navy)
ufs = ["SP", "RJ", "MG", "PR", "RS", "Outros"]
for idx, uf in enumerate(ufs, 10):
    dash.cell(idx, 4, uf)
    formula = f'=COUNTIF(IMPORTACAO_CSV!$I$2:$I$13,"{uf}")' if uf != "Outros" else '=COUNTA(IMPORTACAO_CSV!$I$2:$I$13)-SUM(E10:E14)'
    dash.cell(idx, 5, formula)
    dash.cell(idx, 4).font = Font(name="Arial", color=white)
    dash.cell(idx, 5).font = Font(name="Arial", bold=True, color=white)
    dash.cell(idx, 4).fill = dash.cell(idx, 5).fill = PatternFill("solid", fgColor=panel)

# Hidden helper data with fake status per imported lead.
dash["I11"], dash["J11"] = "Lead", "Status fictício"
for idx, (row, status) in enumerate(zip(rows, statuses), 12):
    dash.cell(idx, 9, row[0])
    dash.cell(idx, 10, status)
dash.column_dimensions["I"].hidden = True
dash.column_dimensions["J"].hidden = True

bar = BarChart()
bar.title = "Distribuição da carteira"
bar.height, bar.width = 7, 12
bar.add_data(Reference(dash, min_col=2, min_row=9, max_row=15), titles_from_data=True)
bar.set_categories(Reference(dash, min_col=1, min_row=10, max_row=15))
bar.legend = None
bar.style = 10
dash.add_chart(bar, "A18")

donut = DoughnutChart()
donut.title = "Origem geográfica"
donut.height, donut.width = 7, 10
donut.add_data(Reference(dash, min_col=5, min_row=9, max_row=15), titles_from_data=True)
donut.set_categories(Reference(dash, min_col=4, min_row=10, max_row=15))
donut.holeSize = 55
dash.add_chart(donut, "E18")

for col in range(1, 9):
    dash.column_dimensions[get_column_letter(col)].width = 16
for row in range(1, 34):
    dash.row_dimensions[row].height = 20
    for col in range(1, 9):
        if not dash.cell(row, col).fill.fill_type:
            dash.cell(row, col).fill = PatternFill("solid", fgColor=navy)

# Guide
guide.column_dimensions["A"].width = 28
guide.column_dimensions["B"].width = 95
guide["A1"], guide["B1"] = "Item", "Orientação"
for cell in guide[1]:
    cell.fill = PatternFill("solid", fgColor=purple)
    cell.font = Font(name="Arial", bold=True, color=white)
guide_rows = [
    ("Arquivo para upload", "Use o CSV prospectos-importacao-demo.csv. A plataforma aceita CSV, não o XLSX."),
    ("Campos obrigatórios", "Nome, E-mail, Empresa e Segmento."),
    ("Campos opcionais", "Origem, Telefone, Cargo, Cidade e Estado."),
    ("O que editar", "Substitua os exemplos da aba IMPORTACAO_CSV pelos dados reais e exporte essa aba como CSV separado por ponto e vírgula."),
    ("Atribuição", "Na versão planejada, todos os leads serão atribuídos automaticamente ao comercial autenticado."),
    ("Segurança", "Importar não inicia campanha nem envia e-mails. O gestor é avisado e ativa a campanha depois."),
    ("Dados", "Todos os nomes, empresas, e-mails e números deste arquivo são fictícios."),
]
for r_idx, (item, text) in enumerate(guide_rows, 2):
    guide.cell(r_idx, 1, item).font = Font(name="Arial", bold=True, color="1E293B")
    guide.cell(r_idx, 2, text).font = Font(name="Arial", color="334155")
    guide.cell(r_idx, 1).fill = guide.cell(r_idx, 2).fill = PatternFill("solid", fgColor="F8FAFC" if r_idx % 2 == 0 else "EEF2FF")
    guide.cell(r_idx, 2).alignment = Alignment(wrap_text=True, vertical="top")
    guide.row_dimensions[r_idx].height = 38

wb.calculation.fullCalcOnLoad = True
wb.calculation.forceFullCalc = True
wb.calculation.calcMode = "auto"
wb.save(XLSX)
print(XLSX)
print(CSV)
